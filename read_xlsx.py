from openpyxl import load_workbook

# Передаем путь к папке с файлами (resources)
workbook = load_workbook('resources/file_example_XLSX_50.xlsx')
sheat = workbook.active
print(sheat.cell(row=6, column=3).value)